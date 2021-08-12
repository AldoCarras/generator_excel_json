from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("BD_EXCEL.xlsx")
sheet = wb.get_sheet_by_name('Hoja 1')
sheet = wb.active

rows = 0
for max_row, row in enumerate(sheet, 1):
    if not all(col.value is None for col in row):
        rows += 1

for i in range(2, rows+1):
    print(sheet.cell(row=i, column=3).value)






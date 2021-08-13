from openpyxl import load_workbook
import json

wb = load_workbook("BD_EXCEL.xlsx")
data = []
for sheet_title in wb.get_sheet_names():
    sheet = wb.get_sheet_by_name(sheet_title)
    for row in sheet.rows:
        fields = {}
        sw = 0
        if row[0].value is None:
            break
        for cell in row:
            if cell.row == 1:
                sw = 1
            else:
                if cell.value is None:
                    fields[sheet.cell(row=1, column=cell.column).value] = ""
                else:
                    fields[sheet.cell(row=1, column=cell.column).value] = cell.value
        if sw == 0:
            data.append({
                "model": sheet_title,
                "fields": fields
            })

with open('data.json', 'w') as file:
    json.dump(data, file, indent=4)






